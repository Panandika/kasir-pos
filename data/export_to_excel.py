"""
Export kasir.db to a single Excel file, one sheet per table, top 100 rows each.
Output: kasir-export.xlsx (same directory as this script)
"""

import sqlite3
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DB_PATH = os.path.join(os.path.dirname(__file__), "kasir.db")
OUT_PATH = os.path.join(os.path.dirname(__file__), "kasir-export.xlsx")
ROW_LIMIT = 100

# FTS internal tables — not readable as plain data
SKIP_TABLES = {
    "products_fts",
    "products_fts_config",
    "products_fts_data",
    "products_fts_docsize",
    "products_fts_idx",
}

HEADER_FILL = PatternFill("solid", fgColor="2B2B2B")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
CELL_FONT   = Font(name="Calibri", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def get_tables(conn):
    cur = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
    )
    all_tables = [row[0] for row in cur.fetchall()]

    # Separate non-empty vs empty so non-empty sheets come first
    non_empty, empty = [], []
    for t in all_tables:
        if t in SKIP_TABLES:
            continue
        count = conn.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0]
        if count > 0:
            non_empty.append((t, count))
        else:
            empty.append((t, 0))

    # Sort non-empty by row count descending for easy navigation
    non_empty.sort(key=lambda x: x[1], reverse=True)
    return non_empty + empty


def safe_sheet_name(name):
    # Excel sheet names: max 31 chars, no special chars
    invalid = r'\/*?:[]\''
    for ch in invalid:
        name = name.replace(ch, "_")
    return name[:31]


def auto_width(ws, col_idx, values):
    max_len = max((len(str(v)) if v is not None else 0 for v in values), default=8)
    max_len = min(max(max_len, 8), 50)  # clamp between 8 and 50
    ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


def write_sheet(ws, conn, table_name):
    cur = conn.execute(f'SELECT * FROM "{table_name}" LIMIT {ROW_LIMIT}')
    columns = [desc[0] for desc in cur.description]
    rows = cur.fetchall()

    # Header row
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT

    # Auto-width per column
    for col_idx, col_name in enumerate(columns, start=1):
        col_values = [col_name] + [row[col_idx - 1] for row in rows]
        auto_width(ws, col_idx, col_values)

    # Freeze header row
    ws.freeze_panes = "A2"

    return len(rows)


def main():
    print(f"Connecting to: {DB_PATH}")
    conn = sqlite3.connect(DB_PATH)

    wb = Workbook()
    wb.remove(wb.active)  # Remove default empty sheet

    tables = get_tables(conn)
    print(f"Found {len(tables)} tables to export\n")

    total_rows = 0
    for table_name, _ in tables:
        sheet_name = safe_sheet_name(table_name)
        ws = wb.create_sheet(title=sheet_name)
        exported = write_sheet(ws, conn, table_name)
        total_rows += exported
        status = f"{exported:>3} rows" if exported > 0 else "  (empty)"
        print(f"  {table_name:<35} {status}")

    conn.close()

    wb.save(OUT_PATH)
    print(f"\nDone. {total_rows} total rows across {len(tables)} sheets.")
    print(f"Saved to: {OUT_PATH}")


if __name__ == "__main__":
    main()
